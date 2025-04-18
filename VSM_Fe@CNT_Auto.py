# VSMで得られたテキストファイルからエクセルを作成するコード

import os
import sys
import pandas as pd
import chardet
import re
import shutil
from openpyxl import load_workbook

# ユーザー入力による連番とファイル数の指定
start_number = int(input("ファイル名の開始番号を入力してください（例:1500）: "))
num_files = int(input("回数を入力してください（例: 16）: "))

# --- ファイルパスの設定 ---
template_excel = "VSM_Fe@CNT_テンプレート.xlsx"    # 元のExcelファイル（変更しない）
text_folder = r"C:\Users\ne4\OneDrive - 三重大学\ドキュメント\三重大\ナノエレクトロニクス研究室\自動処理\テキストファイル"
save_directory = r"C:\Users\ne4\OneDrive - 三重大学\ドキュメント\三重大\ナノエレクトロニクス研究室\自動処理\エクセルファイル"

for number in range(start_number, start_number + num_files):
    # テキストファイルと新規Excelファイルのパスを作成
    text_file = os.path.join(text_folder, f"F{number}_perp.txt")    # テキストファイル
    new_excel_filename = f"F{number}.xlsx"    # 新しく保存するExcelファイル
    new_excel = os.path.join(save_directory, new_excel_filename)

    # テキストファイルが存在するかチェック（存在しなければエラーメッセージを表示して終了）
    if not os.path.exists(text_file):
        print(f"エラー: テキストファイル {text_file} が見つかりません。")
        sys.exit(1)

    # --- 文字コードの自動判定 ---
    with open(text_file, "rb") as f:
        result = chardet.detect(f.read(10000))
        encoding = result["encoding"]

    # --- テキストファイルからデータ抽出 ---
    cleaned_data = []   # 2列の数値データ格納用リスト
    hc_value = None     # "Hc" に続く数値を格納

    with open(text_file, "r", encoding=encoding) as file:
        for line in file:
            # --- Hc の値を抽出 ---
            if hc_value is None and "Hc" in line:
                # "Hc" の後に続く数値（整数、小数、科学表記対応）を抽出
                match = re.search(r'Hc\s*[:=]?\s*([-+]?[0-9]*\.?[0-9]+(?:[eE][-+]?[0-9]+)?)', line)
                if match:
                    try:
                        hc_value = float(match.group(1).replace("E", "e"))
                    except ValueError:
                        pass  # 数値変換できなかった場合は無視
            
            # --- 2列の数値データ (Oe, emu) の抽出 ---
            # 不要な行（ヘッダーやその他のメタ情報）をスキップ
            if line.strip() and not any(x in line for x in ["Model", "File Name", "測定日時", "測定番号", "試料名", "補正", "Image File", "BG File"]):
                parts = line.split("\t")  # タブ区切りで分割
                if len(parts) == 2:  # 2列データの場合のみ処理
                    try:
                        col1 = float(parts[0].replace("E", "e"))
                        col2 = float(parts[1].replace("E", "e"))
                        cleaned_data.append([col1, col2])
                    except ValueError:
                        continue  # 数値に変換できない行はスキップ

    # --- DataFrame の作成 ---
    df = pd.DataFrame(cleaned_data, columns=["Oe", "emu"])

    # --- VSM_Fe@CNT.xlsx をコピーして F~~~~.xlsx として保存 ---
    shutil.copy(template_excel, new_excel)

    # --- コピーした Excel ファイルを openpyxl で開く ---
    wb = load_workbook(new_excel)
    ws = wb.active  # ここではアクティブなシートに貼り付ける前提

    # --- セル Q4 に Hc の値を貼り付け ---
    if hc_value is not None:
        ws["Q4"] = hc_value
    else:
        print("Hc の値がテキスト内に見つかりませんでした。")

    # --- DataFrame のデータをシートの I, J 列に貼り付け ---
    # ここではヘッダーを I1, J1に配置します
    start_row = 1  # ヘッダー行の開始位置
    start_col = 9  # I列 (openpyxlでは1がA列)

    # ヘッダーの書き込み
    for j, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j)
        cell.value = col_name


    # データ行の書き込み（ヘッダーの下の行から開始）
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row):
            cell = ws.cell(row=i, column=start_col + j)
            cell.value = value  # ← 値だけ代入する。フォーマットはそのまま保持。


    # --- 変更内容を保存 ---
    wb.save(new_excel)

    print(f"{template_excel} の内容は変更せず、コピー先の {new_excel} にデータを貼り付けました。")
